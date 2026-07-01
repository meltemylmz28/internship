import csv
import os
import re
import shutil
import subprocess
from pathlib import Path
import cv2
import numpy as np
import openpyxl
import pandas as pd
import xlrd
from docx import Document
from lxml import etree
from PIL import Image
from pypdf import PdfReader
import pytesseract
import pytesseract

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\tesseract.exe'
 
def process_excel_file(file_path):
    """
    Yüklenen dosyayı otomatik olarak analiz eder.
    Excel, CSV, XML, PDF, görsel, Word ve TXT formatlarını destekler.
    Her sınıfın dersleri ve başlıkları dinamik olarak ayrıştırılır.
    """
    try:
        extension = Path(file_path).suffix.lower()
        if extension == '.csv':
            df = _read_csv(file_path)
        elif extension in {'.xls', '.xlsx'}:
            students_data = _read_excel(file_path)
            if students_data:
                return {"students": students_data}
            return {"students": [], "error": "Excel dosyası okunamadı veya içeriği ayrıştırılamadı."}
        elif extension == '.xml':
            df = _read_xml(file_path)
        elif extension == '.pdf':
            df = _read_pdf(file_path)
        elif extension in {'.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tif', '.tiff'}:
            df = _read_image(file_path)
        elif extension in {'.doc', '.docx'}:
            df = _read_doc(file_path)
        elif extension in {'.txt', '.rtf'}:
            df = _read_text(file_path)
        else:
            return {"error": f"Desteklenmeyen dosya türü: {extension}", "students": []}
    except Exception as e:
        return {"error": f"Dosya okunamadı: {str(e)}", "students": []}

    if df is None or df.empty:
        return {"error": "Dosya içeriği boş.", "students": []}

    students_data = _extract_students(df)

    if not students_data:
        return {
            "students": [],
            "error": "Belirli bir öğrenci yapısı ayrıştırılamadı ancak dosya başarıyla yüklendi.",
            "raw_data": df.fillna('-').values.tolist()[:10]
        }

    return {"students": students_data}


def _read_csv(file_path):
    for delimiter in [',', ';', '\t', '|']:
        try:
            with open(file_path, 'r', encoding='utf-8-sig', errors='ignore', newline='') as handle:
                rows = list(csv.reader(handle, delimiter=delimiter))
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            if rows and (len(rows) > 1 or any(len(row) > 1 for row in rows)):
                return _rows_to_dataframe(rows)
        except Exception:
            continue

    lines = Path(file_path).read_text(encoding='utf-8-sig', errors='ignore').splitlines()
    rows = []
    for line in lines:
        if not line.strip():
            continue
        rows.append([cell.strip() for cell in line.split(',') if cell.strip()])
    return _rows_to_dataframe(rows)


def _read_excel(file_path):
    extension = Path(file_path).suffix.lower()

    try:
        df = pd.read_excel(file_path, header=None, dtype=str, keep_default_na=False, engine='openpyxl' if extension == '.xlsx' else 'xlrd')
        df = df.fillna('').astype(str)
        students_data = _extract_students(df)
        if students_data:
            return students_data
    except Exception:
        pass

    try:
        if extension == '.xls':
            students = xls_oku(file_path)
        else:
            students = xlsx_oku(file_path)
        if students:
            return students
    except Exception:
        pass

    return []


def _normalize_students(student_records):
    normalized = []
    for entry in student_records:
        name = entry.get('ogrenci_adi') or entry.get('name') or entry.get('ogrenci') or 'Öğrenci'
        grades = entry.get('notlar') or entry.get('grades') or {}
        number = entry.get('okul_no') or entry.get('number') or entry.get('no') or entry.get('sira') or '-'
        ortalama = entry.get('ortalama')
        if ortalama is not None and 'Genel Sonuç/Değerlendirme' not in grades:
            grades['Genel Sonuç/Değerlendirme'] = ortalama
        normalized.append({
            'name': name,
            'number': number,
            'grades': grades
        })
    return normalized


def xls_oku(dosya_yolu):
    wb = xlrd.open_workbook(dosya_yolu)
    konsolide_veriler = {}

    ders_haritasi = {
        'T': 'TÜRKÇE', 'M': 'MATEMATİK', 'HB': 'HAYAT BİLGİSİ',
        'BEO': 'BEDEN EĞİTİMİ VE OYUN', 'MÜ': 'MÜZİK', 'GS': 'GÖRSEL SANATLAR'
    }

    for sheet_name in wb.sheet_names():
        if 'bilgi' in sheet_name.lower() or 'not' in sheet_name.lower():
            continue

        sheet = wb.sheet_by_name(sheet_name)
        temiz_kod = re.sub(r'[^a-zA-ZğüşıöçĞÜŞİÖÇ]', '', sheet_name).upper().strip()
        ders_adi = ders_haritasi.get(temiz_kod, f'DERS ({sheet_name})')

        baslik_satiri = 3
        for r in range(min(sheet.nrows, 6)):
            if sheet.ncols > 2:
                h_val = str(sheet.cell_value(r, 2) or '').upper()
                if 'ADI' in h_val or 'SOYADI' in h_val or 'SIRA' in str(sheet.cell_value(r, 0) or '').upper():
                    baslik_satiri = r
                    break

        puan_sutunlari = []
        if sheet.nrows > baslik_satiri:
            for col in range(3, sheet.ncols):
                b_degeri = str(sheet.cell_value(baslik_satiri, col) or '').upper()
                if 'ORTALAMA' in b_degeri or 'SONUÇ' in b_degeri or b_degeri == '':
                    continue
                puan_sutunlari.append(col)

            for r in range(baslik_satiri + 1, sheet.nrows):
                ogrenci_adi = sheet.cell_value(r, 2)
                okul_no_val = sheet.cell_value(r, 1)
                if not ogrenci_adi:
                    continue

                ogrenci_adi_str = str(ogrenci_adi).strip().upper()
                if any(k in ogrenci_adi_str for k in ['TOPLAM', 'SINIF ÖĞRETMENİ', 'OKUL MÜDÜRÜ', 'GRAFİK', 'MÜDÜR']):
                    continue

                if okul_no_val is None or str(okul_no_val).strip() == '' or 'OKUL' in str(okul_no_val).upper():
                    continue

                toplam_puan = 0
                gecerli_puan_sayisi = 0
                grades = {}
                for col in puan_sutunlari:
                    hucre_degeri = sheet.cell_value(r, col)
                    if hucre_degeri is not None and str(hucre_degeri).strip() != '':
                        try:
                            puan_float = float(hucre_degeri)
                            toplam_puan += puan_float
                            gecerli_puan_sayisi += 1
                            grades[f'Ders {col}'] = puan_float
                        except ValueError:
                            continue

                if gecerli_puan_sayisi > 0:
                    ham_ortalama = toplam_puan / gecerli_puan_sayisi
                    yuzluk_not = (ham_ortalama / 4.0) * 100 if ham_ortalama <= 4.0 else ham_ortalama
                    if ogrenci_adi_str not in konsolide_veriler:
                        konsolide_veriler[ogrenci_adi_str] = {}
                    konsolide_veriler[ogrenci_adi_str][ders_adi] = round(yuzluk_not, 2)

    ogrenci_listesi = []
    for ogr_ad, dersler in konsolide_veriler.items():
        if not dersler:
            continue
        genel_ort = sum(dersler.values()) / len(dersler)
        ogrenci_listesi.append({
            'ogrenci_adi': ogr_ad,
            'notlar': dersler,
            'ortalama': round(genel_ort, 2)
        })
    return _normalize_students(ogrenci_listesi)


def _find_header_row(sheet, search_columns, keywords, max_scan_row=12):
    for r in range(1, min(max_scan_row, sheet.max_row) + 1):
        row_values = [str(sheet.cell(row=r, column=c).value or '').upper() for c in search_columns]
        joined = ' '.join(row_values)
        if any(keyword in joined for keyword in keywords):
            return r
    return None


def _normalize_row_values(values):
    return [str(v).strip() if v is not None else '' for v in values]


def xlsx_oku(dosya_yolu):
    wb = openpyxl.load_workbook(dosya_yolu, data_only=True)
    if len(wb.sheetnames) == 1:
        sheet = wb.active
        ogrenci_listesi = []
        dersler = []
        BASLANGIC_SUTUNU = 2
        BASLIK_SATIRI = _find_header_row(sheet, range(1, min(8, sheet.max_column) + 1), ['ÖĞRENCİ ADI', 'AD SOYAD', 'ÖĞRENCİNİN ADI'], max_scan_row=10) or 5

        for col in range(BASLANGIC_SUTUNU, sheet.max_column + 1):
            ders_adi = sheet.cell(row=BASLIK_SATIRI, column=col).value
            if ders_adi and str(ders_adi).strip():
                dersler.append(str(ders_adi).strip())

        start_row = BASLIK_SATIRI + 1
        for row in range(start_row, sheet.max_row + 1):
            ogrenci_girdi = sheet.cell(row=row, column=1).value
            if ogrenci_girdi is None or str(ogrenci_girdi).strip() == '':
                continue
            ogrenci_adi = str(ogrenci_girdi).replace('Öğrenci Adı:', '').replace('ÖĞRENCİ ADI:', '').strip().upper()
            if not ogrenci_adi:
                continue

            notlar = {}
            toplam_not = 0
            ders_sayisi = 0
            any_grade = False

            for idx, ders in enumerate(dersler):
                not_degeri = sheet.cell(row=row, column=idx + BASLANGIC_SUTUNU).value
                if not_degeri is not None and str(not_degeri).strip() != '':
                    try:
                        not_float = float(not_degeri)
                        notlar[ders] = round(not_float, 2)
                        toplam_not += not_float
                        ders_sayisi += 1
                        any_grade = True
                    except ValueError:
                        continue

            if not any_grade:
                continue

            ortalama = toplam_not / ders_sayisi if ders_sayisi > 0 else 0
            ogrenci_listesi.append({
                'ogrenci_adi': ogrenci_adi,
                'notlar': notlar,
                'ortalama': round(ortalama, 2)
            })
        return _normalize_students(ogrenci_listesi)

    konsolide_veriler = {}
    ders_haritasi = {
        'T': 'TÜRKÇE', 'M': 'MATEMATİK', 'HB': 'HAYAT BİLGİSİ',
        'BEO': 'BEDEN EĞİTİMİ VE OYUN', 'MÜ': 'MÜZİK', 'GS': 'GÖRSEL SANATLAR'
    }

    for sheet_name in wb.sheetnames:
        if 'bilgi' in sheet_name.lower():
            continue

        sheet = wb[sheet_name]
        temiz_kod = re.sub(r'[^a-zA-ZğüşıöçĞÜŞİÖÇ]', '', sheet_name).upper().strip()
        ders_adi = ders_haritasi.get(temiz_kod, f'DERS ({sheet_name})')

        baslik_satiri = 4
        for r in range(1, 6):
            h_val = str(sheet.cell(row=r, column=3).value or '').upper()
            if 'ADI' in h_val or 'SOYADI' in h_val:
                baslik_satiri = r
                break

        puan_sutunlari = []
        for col in range(4, sheet.max_column + 1):
            b_degeri = str(sheet.cell(row=baslik_satiri, column=col).value or '').upper()
            if 'ORTALAMA' in b_degeri or 'SONUÇ' in b_degeri or b_degeri == '':
                continue
            puan_sutunlari.append(col)

        for r in range(baslik_satiri + 1, sheet.max_row + 1):
            ogrenci_adi = sheet.cell(row=r, column=3).value
            if ogrenci_adi is None or str(ogrenci_adi).strip() == '':
                continue
            ogrenci_adi_str = str(ogrenci_adi).strip().upper()
            if any(k in ogrenci_adi_str for k in ['TOPLAM', 'ORTALAMA', 'SINIF', 'ÖĞRENCİ', 'DERS']):
                continue

            toplam_puan = 0
            gecerli_puan_sayisi = 0
            any_grade = False

            for col in puan_sutunlari:
                hucre_degeri = sheet.cell(row=r, column=col).value
                if hucre_degeri is not None and str(hucre_degeri).strip() != '':
                    try:
                        puan_float = float(hucre_degeri)
                        toplam_puan += puan_float
                        gecerli_puan_sayisi += 1
                        any_grade = True
                    except ValueError:
                        continue

            if not any_grade:
                continue

            ham_ortalama = toplam_puan / gecerli_puan_sayisi
            yuzluk_not = (ham_ortalama / 4.0) * 100 if ham_ortalama <= 4.0 else ham_ortalama
            if ogrenci_adi_str not in konsolide_veriler:
                konsolide_veriler[ogrenci_adi_str] = {}
            konsolide_veriler[ogrenci_adi_str][ders_adi] = round(yuzluk_not, 2)

    ogrenci_listesi = []
    for ogr_ad, dersler in konsolide_veriler.items():
        if not dersler:
            continue
        genel_ort = sum(dersler.values()) / len(dersler)
        ogrenci_listesi.append({
            'ogrenci_adi': ogr_ad,
            'notlar': dersler,
            'ortalama': round(genel_ort, 2)
        })
    return _normalize_students(ogrenci_listesi)


def _read_xml(file_path):
    tree = etree.parse(file_path)
    root = tree.getroot()
    rows = []
    for element in root.iter():
        if element.tag.lower() in {'row', 'record', 'tr'}:
            values = []
            for child in element.iterchildren():
                values.append(_clean_text(child.text))
            if values:
                rows.append(values)
    if rows:
        return _rows_to_dataframe(rows)
    return _rows_to_dataframe([[text] for text in etree.tostring(root, encoding='unicode').splitlines() if text.strip()])


def _read_pdf(file_path):
    reader = PdfReader(file_path)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    return _rows_to_dataframe(_split_text_rows(text))


def _find_tesseract_cmd():
    # PATH üzerinde tesseract aranır
    tesseract_cmd = shutil.which('tesseract')
    if tesseract_cmd:
        return tesseract_cmd

    # Windows için yaygın yükleme yolları
    common_paths = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]
    for path in common_paths:
        if Path(path).exists():
            return path

    return None


def _read_image(file_path):
    try:
        image = Image.open(file_path)
        tesseract_cmd = _find_tesseract_cmd()
        if not tesseract_cmd:
            raise RuntimeError(
                'tesseract yüklü değil veya PATH içinde bulunamadı. ' \
                'Lütfen Tesseract OCR programını yükleyin ve PATH değişkenine ekleyin.'
            )
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
        text = pytesseract.image_to_string(image)
        return _rows_to_dataframe(_split_text_rows(text))
    except Exception as exc:
        raise RuntimeError(f"Görsel OCR işlenemedi: {exc}") from exc


def _read_doc(file_path):
    extension = Path(file_path).suffix.lower()
    if extension == '.docx':
        document = Document(file_path)
        text = "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text)
        return _rows_to_dataframe(_split_text_rows(text))

    antiword_path = shutil.which('antiword')
    if antiword_path:
        result = subprocess.run([antiword_path, file_path], capture_output=True, text=True, check=False)
        if result.returncode == 0 and result.stdout.strip():
            return _rows_to_dataframe(_split_text_rows(result.stdout))

    raise RuntimeError('DOC dosyası desteklenmiyor. antiword kurulu değil.')


def _read_text(file_path):
    text = Path(file_path).read_text(encoding='utf-8', errors='ignore')
    return _rows_to_dataframe(_split_text_rows(text))


def _rows_to_dataframe(rows):
    if not rows:
        return pd.DataFrame()
    width = max(len(row) for row in rows)
    normalized_rows = []
    for row in rows:
        padded = [ _clean_text(value) for value in row ]
        if len(padded) < width:
            padded.extend([''] * (width - len(padded)))
        normalized_rows.append(padded)
    return pd.DataFrame(normalized_rows, dtype=str)


def _split_text_rows(text):
    if not text:
        return []
    lines = []
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if '\t' in line:
            parts = [part.strip() for part in line.split('\t') if part.strip()]
        else:
            parts = [part.strip() for part in re.split(r'\s{2,}|\s*;\s*', stripped) if part.strip()]
        lines.append(parts if parts else [stripped])
    return lines


def _extract_students(df):
    cleaned_df = _clean_dataframe(df)
    if cleaned_df.empty:
        return []

    students = []
    direct_name = _find_direct_student_name(cleaned_df)
    if direct_name:
        parsed = _parse_single_student_format(cleaned_df, direct_name)
        if parsed:
            return parsed

    header_row_idx = _find_header_row(cleaned_df)
    if header_row_idx is None:
        return _parse_rows_without_header(cleaned_df)

    header_row = [_clean_text(value) for value in cleaned_df.iloc[header_row_idx].tolist()]
    name_col_idx = _find_column_index(header_row, ['ad', 'soyad', 'isim', 'name', 'öğrenci'])
    no_col_idx = _find_column_index(header_row, ['no', 'numara', 'okul', 'sıra'])
    result_col_idx = _find_column_index(header_row, ['sonuç', 'ortalama', 'değerlendirme', 'durum', 'genel'])

    grade_columns = []
    for idx, title in enumerate(header_row):
        if idx in {name_col_idx, no_col_idx, result_col_idx}:
            continue
        if title and _looks_like_subject(title):
            grade_columns.append((idx, title))

    if not grade_columns:
        grade_columns = [(idx, title) for idx, title in enumerate(header_row) if idx not in {name_col_idx, no_col_idx, result_col_idx} and title]

    if name_col_idx is None and header_row:
        name_col_idx = 0

    for row_idx in range(header_row_idx + 1, len(cleaned_df)):
        row = [_clean_text(value) for value in cleaned_df.iloc[row_idx].tolist()]
        if not any(row):
            continue

        name_value = row[name_col_idx] if name_col_idx is not None and name_col_idx < len(row) else ''
        if not name_value or _looks_like_numeric(name_value) or name_value.lower() in {'ad', 'soyad', 'isim', 'öğrenci'}:
            continue

        no_value = row[no_col_idx] if no_col_idx is not None and no_col_idx < len(row) else '-'
        grades = {}
        for col_idx, course_name in grade_columns:
            if col_idx < len(row):
                value = row[col_idx]
                if value and not _looks_like_numeric_only(course_name):
                    grades[course_name] = value

        if result_col_idx is not None and result_col_idx < len(row):
            result_value = row[result_col_idx]
            if result_value:
                grades['Genel Sonuç/Değerlendirme'] = result_value

        if not grades:
            continue

        students.append({
            'name': name_value,
            'number': no_value,
            'grades': grades
        })

    return students


def _parse_rows_without_header(df):
    students = []
    for _, row in df.iterrows():
        values = [_clean_text(value) for value in row.tolist()]
        if not any(values):
            continue

        first_value = values[0]
        if not first_value or _looks_like_numeric(first_value):
            continue

        grades = {}
        for idx, val in enumerate(values[1:], start=1):
            if val and not _looks_like_numeric_only(values[0]):
                grades[f'Sütun {idx + 1}'] = val

        if grades:
            students.append({
                'name': first_value,
                'number': values[1] if len(values) > 1 else '-',
                'grades': grades
            })
    return students


def _parse_single_student_format(df, student_name):
    start_idx = None
    for idx, row in df.iterrows():
        joined = ' '.join(_clean_text(value) for value in row.tolist() if _clean_text(value))
        if any(keyword in joined.lower() for keyword in ['öğrenci adı', 'student name', 'öğrencinin adı soyadı']):
            start_idx = idx
            if student_name is None:
                student_name = _extract_name_from_row(joined)
            break

    if start_idx is None:
        return []

    for candidate_idx in range(start_idx + 1, min(len(df), start_idx + 12)):
        candidate_row = [_clean_text(value) for value in df.iloc[candidate_idx].tolist()]
        if not any(candidate_row):
            continue

        lower_joined = ' '.join(candidate_row).lower()
        if 'dersler' in lower_joined or '1. dönem' in lower_joined or 'haftalık ders saati' in lower_joined:
            header_row = candidate_row
            data_rows = []
            for next_idx in range(candidate_idx + 1, len(df)):
                next_row = [_clean_text(value) for value in df.iloc[next_idx].tolist()]
                if not any(next_row):
                    continue
                if _looks_like_subject_row(next_row):
                    data_rows.append(next_row)
                else:
                    break

            grades = {}
            if data_rows:
                keys = [title for title in header_row if title]
                for row in data_rows:
                    subject = row[0] if len(row) > 0 else ''
                    if not subject or _looks_like_numeric_only(subject):
                        continue
                    values = [cell for cell in row[1:] if cell]
                    if values:
                        grades[subject] = values[0]
                if grades:
                    return [{"name": student_name or 'Öğrenci', "number": '-', "grades": grades}]

        if _looks_like_subject_row(candidate_row):
            values = [_clean_text(value) for value in df.iloc[candidate_idx + 1].tolist()] if candidate_idx + 1 < len(df) else []
            grades = {}
            for title, value in zip(candidate_row, values):
                if title and value and not _looks_like_numeric_only(title):
                    grades[title] = value
            if grades:
                return [{"name": student_name or 'Öğrenci', "number": '-', "grades": grades}]
            return [{"name": student_name or 'Öğrenci', "number": '-', "grades": {}}]

    return [{"name": student_name or 'Öğrenci', "number": '-', "grades": {}}]


def _find_direct_student_name(df):
    for _, row in df.iterrows():
        joined = ' '.join(_clean_text(value) for value in row.tolist() if _clean_text(value))
        if any(keyword in joined.lower() for keyword in ['öğrenci adı', 'student name', 'ad soyad', 'öğrencinin adı soyadı']):
            parts = [part.strip() for part in re.split(r'[:\-]', joined, maxsplit=1)]
            if len(parts) > 1:
                return parts[-1].replace(',', '').strip()
    return None


def _extract_name_from_row(row_text):
    if ':' in row_text:
        return row_text.split(':', 1)[-1].strip()
    if '-' in row_text:
        return row_text.split('-', 1)[-1].strip()
    return None


def _find_header_row(df):
    for idx, row in df.iterrows():
        values = [_clean_text(value) for value in row.tolist() if _clean_text(value)]
        if len(values) < 2:
            continue
        if any(keyword in ' '.join(values).lower() for keyword in ['öğrenci', 'student', 'ad soyad', 'isim', 'numara', 'no', 'ders', 'kazanım', 'puan', 'not', 'ortalama', 'sonuç', 'değerlendirme']):
            return idx
        if _looks_like_subject_row(values):
            return idx
    return None


def _find_column_index(header_row, keywords):
    for idx, title in enumerate(header_row):
        lowered = _clean_text(title).lower()
        if any(keyword in lowered for keyword in keywords):
            return idx
    return None


def _looks_like_subject_row(values):
    if len(values) < 2:
        return False
    text_like_count = 0
    for value in values:
        lowered = value.lower()
        if re.search(r'[a-zçğışöü]', lowered) or not re.fullmatch(r'[\d\.,%/\-]+', lowered):
            text_like_count += 1
    return text_like_count >= max(2, len(values) // 2)


def _looks_like_subject(title):
    lowered = title.lower()
    if any(keyword in lowered for keyword in ['öğrenci', 'student', 'ad', 'soyad', 'isim', 'no', 'numara', 'sıra', 'sonuç', 'ortalama', 'değerlendirme', 'genel', 'durum']):
        return False
    return bool(re.search(r'[a-zçğışöü]', lowered))


def _looks_like_numeric(value):
    return bool(re.fullmatch(r'[\d\.,%/\-]+', value))


def _looks_like_numeric_only(value):
    return bool(re.fullmatch(r'[\d\.,%/\-]+', value))


def _clean_dataframe(df):
    rows = []
    for _, row in df.iterrows():
        cleaned = [_clean_text(value) for value in row.tolist()]
        if any(cleaned):
            rows.append(cleaned)

    if not rows:
        return pd.DataFrame()

    max_len = max(len(row) for row in rows)
    normalized_rows = []
    for row in rows:
        if len(row) < max_len:
            row = row + [''] * (max_len - len(row))
        normalized_rows.append(row)

    return pd.DataFrame(normalized_rows, dtype=str)


def _clean_text(value):
    if value is None:
        return ''
    if pd.isna(value):
        return ''
    if isinstance(value, (int, float)) and pd.isna(value):
        return ''
    text = str(value).strip()
    return re.sub(r'\s+', ' ', text)